# src/summary.py
from builtins import enumerate, float, print,any
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from config.column_names import *

def create_multi_metric_forecast_summary(df, metrics, output_file="outputs/Multi_Metric_Forecast_Summary.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Summary"

    stats_measures = ["Median", "10th Percentile", "90th Percentile"]

    current_row = 1
    tickers = df[col_ticker].unique()

    for ticker in tickers:
        ticker_data = df[df[col_ticker] == ticker].copy()

        ws.cell(row=current_row, column=1, value=f"{ticker} FORECAST PANEL")
        current_row += 1

        display_cols = [
            col_ticker, col_broker_name, col_analyst_name,
            *metrics
        ]
        valid_cols = [col for col in display_cols if col in df.columns]

        for col_idx, header in enumerate(valid_cols, 1):
            ws.cell(row=current_row, column=col_idx, value=header)
        current_row += 1

        start_data_row = current_row
        for _, row in ticker_data.iterrows():
            for col_idx, header in enumerate(valid_cols, 1):
                value = row.get(header, None)
                if pd.isna(value):
                    value = None
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        end_data_row = current_row - 1

        current_row += 1

        ws.cell(row=current_row, column=1, value=f"Summary Statistics - {ticker}")
        current_row += 1

        ws.cell(row=current_row, column=1, value="Statistic")
        for col_idx, metric in enumerate(metrics, 2):
            ws.cell(row=current_row, column=col_idx, value=metric)
        current_row += 1

        for stat in stats_measures:
            ws.cell(row=current_row, column=1, value=stat)
            for col_idx, metric in enumerate(metrics, 2):
                if metric not in valid_cols:
                    continue
                metric_col_idx = valid_cols.index(metric) + 1
                metric_col_letter = get_column_letter(metric_col_idx)
                data_range = f"{metric_col_letter}{start_data_row}:{metric_col_letter}{end_data_row}"

                values = ticker_data[metric].dropna().astype(float)

                if values.empty:
                    ws.cell(row=current_row, column=col_idx, value=None)
                else:
                    if stat == "Median":
                        formula = f"=MEDIAN({data_range})"
                        ws.cell(row=current_row, column=col_idx, value=formula)
                    else:
                        percentile = 10 if stat == "10th Percentile" else 90
                        value = np.percentile(values[values != 0], percentile)
                        ws.cell(row=current_row, column=col_idx, value=value)

                    if any(m in metric for m in ["Margin", "Dividend Yield"]):
                        ws.cell(row=current_row, column=col_idx).number_format = '0.0%'

            current_row += 1

        current_row += 2

    wb.save(output_file)
    print(f"Multi-metric forecast summary saved to {output_file}")
